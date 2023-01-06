import openpyxl #处理xlsx和xls，不再使用xlwt和xlrd读取，推荐用openpyxl了

# 打开excel
excel = openpyxl.load_workbook('/Users/zoujunping/Python/代码run后保存路径/爬虫兼职1.xlsx')  # 有路径应带上路径
# 使用指定工作表
sheet = excel.active  # 当前激活的工作表
# sheet = excel.get_sheet_by_name('Sheet1')
# 读取所有数据
print(list(sheet.values))  # sheet.values 生成器
print('最大列数:' + str(sheet.max_column))  # 最大列数 4
print('最大行数:' + str(sheet.max_row))  # 最大行数 102



# 按行读取
for row in sheet.iter_rows(min_row=1, min_col=1, max_col=3, max_row=3):
    print(f'打印行\n{row}')

# 读取标题行
for row in sheet.iter_rows(max_row=1):
    title_row = [cell.value for cell in row]
print(f'打印标题行:\n{title_row}')
# 读取标题行以外数据
for row in sheet.iter_rows(min_row=2):
    row_data = [cell.value for cell in row]
    print(row_data)

# 读取单元格数据
print(sheet['C3'].value) #麻省理工学院
print(sheet.cell(1,1).value)  # 索引从1开始

# 写入单元格
sheet['F2'] = 'PASS'
result_col = title_row.index(5)+1  # 'result'所在的列号
sheet.cell(3, result_col).value = 'PASSSD'
# 整行写入
new_row = ['post-xml接口', 'post', 'https://httpbin.org/post']
sheet.append(new_row)
# 保存文件，也可覆盖原文件
excel.save("/Users/zoujunping/Python/代码run后保存路径/apis2.xlsx")




# ----->写入文件
